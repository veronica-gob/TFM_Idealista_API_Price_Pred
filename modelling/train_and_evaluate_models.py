import os
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import KFold
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error, mean_absolute_percentage_error
from openpyxl import load_workbook
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from xgboost import XGBRegressor

class ModelTrainer:
    """
    A class for training multiple regression models and evaluating their performance.
    
    The ModelTrainer class provides methods to train and evaluate different regression models, 
    calculate evaluation metrics (R², RMSE, MAE, MAPE, ME, MPE), and save the results to an Excel file.
    It supports multiple models including Linear Regression, Decision Trees, Random Forest, 
    Gradient Boosting, and XGBoost.
    
    Parameters
    ----------
    metrics_file_path : str
        Path to the Excel file where metrics will be saved.
    """
    def __init__(self, metrics_file_path: str):
        """
        Initializes the ModelTrainer class with the provided metrics file path.
        
        Parameters
        ----------
        metrics_file_path : str
            Path to the Excel file where metrics will be saved.
        """
        self.metrics_file_path = metrics_file_path

    def calculate_metrics(self, y_true, y_pred):
        """
        Calculate various regression metrics for model evaluation.
        
        This method calculates R², RMSE, MAE, MAPE, ME, and MPE between the true values
        and the predicted values.

        Parameters
        ----------
        y_true : array-like
            True target values.
        y_pred : array-like
            Predicted target values.
        
        Returns
        -------
        tuple
            A tuple containing the metrics (R², RMSE, MAE, MAPE, ME, MPE).
        """
        r2 = r2_score(y_true, y_pred)
        rmse = np.sqrt(mean_squared_error(y_true, y_pred))
        mae = mean_absolute_error(y_true, y_pred)
        mape = mean_absolute_percentage_error(y_true, y_pred)
        me = np.mean(y_pred - y_true)
        mpe = np.mean((y_pred - y_true) / y_true) 
        return r2, rmse, mae, mape, me, mpe

    def save_metrics_to_excel(self, metrics: dict) -> None:
        """
        Save the evaluation metrics to an Excel file.
        
        This method appends the evaluation metrics to an existing Excel file or creates a 
        new file if it doesn't exist. The metrics are stored in the 'Metrics' sheet of the file.
        
        Parameters
        ----------
        metrics : dict
            A dictionary containing the evaluation metrics.
        
        Returns
        -------
        None
            This method does not return any value. It saves the metrics to an Excel file.
        """
        df_metrics = pd.DataFrame([metrics])
        
        if os.path.exists(self.metrics_file_path):
            wb = load_workbook(self.metrics_file_path)
            sheet = wb['Metrics']
            next_row = sheet.max_row + 1
            
            for r_idx, row in enumerate(df_metrics.values, next_row):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
            wb.save(self.metrics_file_path)
        else:
            with pd.ExcelWriter(self.metrics_file_path, engine='openpyxl') as writer:
                df_metrics.to_excel(writer, index=False, sheet_name='Metrics')

    def run_linear_regression(self, X, y):
        """
        Train and evaluate other regression models (e.g., Decision Tree, Random Forest, Gradient Boosting, XGBoost).
        
        This method trains and evaluates various models provided in the `models_params` dictionary.
        It uses cross-validation and calculates evaluation metrics for both training and test sets. 
        The results are saved to an Excel file.
        
        Parameters
        ----------
        X : pandas.DataFrame
            Feature matrix (independent variables).
        y : pandas.Series
            Target vector (dependent variable).
        models_params : dict
            A dictionary containing model configurations, including the model object and its parameters.
        
        Returns
        -------
        list
            A list of DataFrames containing the evaluation results for all models.
        """
        # Scale features and target variable
        scaler_features = StandardScaler()
        X_scaled = scaler_features.fit_transform(X)
        scaler_target = StandardScaler()
        y_scaled = scaler_target.fit_transform(y.values.reshape(-1, 1))

        model = LinearRegression()
        cv = KFold(n_splits=5, shuffle=True, random_state=42)
        
        # Store metrics
        train_metrics, test_metrics = [], []
        
        for train_idx, test_idx in cv.split(X_scaled, y_scaled):
            X_train, X_test = X_scaled[train_idx], X_scaled[test_idx]
            y_train, y_test = y_scaled[train_idx], y_scaled[test_idx]
            
            # Train the model
            model.fit(X_train, y_train)
            
            # Predictions
            y_pred_train = model.predict(X_train)
            y_pred_test = model.predict(X_test)

            # Inverse transform predictions
            y_train_original = scaler_target.inverse_transform(y_train.reshape(-1, 1)).flatten()
            y_pred_train_original = scaler_target.inverse_transform(y_pred_train.reshape(-1, 1)).flatten()
            y_test_original = scaler_target.inverse_transform(y_test.reshape(-1, 1)).flatten()
            y_pred_test_original = scaler_target.inverse_transform(y_pred_test.reshape(-1, 1)).flatten()

            # Calculate and append metrics
            train_metrics.append(self.calculate_metrics(y_train_original, y_pred_train_original))
            test_metrics.append(self.calculate_metrics(y_test_original, y_pred_test_original))

        # Compute mean metrics for cross-validation
        mean_train_metrics = np.mean(train_metrics, axis=0)
        mean_test_metrics = np.mean(test_metrics, axis=0)

        model_params = str(model.get_params())

        results = pd.DataFrame({
            'date': [pd.Timestamp.today().strftime('%Y-%m-%d')],
            'model_type': ['Linear Regression'],
            'predictor_variables': [', '.join(X.columns)],  
            'model_parameters': [model_params],  
            'train_r2_mean': [mean_train_metrics[0]],
            'train_rmse_mean': [mean_train_metrics[1]],
            'train_mae_mean': [mean_train_metrics[2]],
            'train_mape_mean': [mean_train_metrics[3]],
            'train_me_mean': [mean_train_metrics[4]],
            'train_mpe_mean': [mean_train_metrics[5]],
            'test_r2_mean': [mean_test_metrics[0]],
            'test_rmse_mean': [mean_test_metrics[1]],
            'test_mae_mean': [mean_test_metrics[2]],
            'test_mape_mean': [mean_test_metrics[3]],
            'test_me_mean': [mean_test_metrics[4]],
            'test_mpe_mean': [mean_test_metrics[5]],
        })
        
        self.save_metrics_to_excel(results.to_dict(orient='records')[0])

    def run_other_models(self, X, y, models_params):
        """
        Train and evaluate other regression models (e.g., Decision Tree, Random Forest, Gradient Boosting, XGBoost).
        
        This method trains and evaluates various models provided in the `models_params` dictionary.
        It uses cross-validation and calculates evaluation metrics for both training and test sets. 
        The results are saved to an Excel file.
        
        Parameters
        ----------
        X : pandas.DataFrame
            Feature matrix (independent variables).
        y : pandas.Series
            Target vector (dependent variable).
        models_params : dict
            A dictionary containing model configurations, including the model object and its parameters.
        
        Returns
        -------
        list
            A list of DataFrames containing the evaluation results for all models.
        """
        cv = KFold(n_splits=5, shuffle=True, random_state=42)
        
        # Store all model results
        all_results = []
        
        for model_key, model_info in models_params.items():
            model = model_info['model']
            model_name = model_info['name']

            train_metrics, test_metrics = [], []

            for train_idx, test_idx in cv.split(X, y):
                X_train, X_test = X.iloc[train_idx], X.iloc[test_idx]
                y_train, y_test = y.iloc[train_idx], y.iloc[test_idx]

                model.fit(X_train, y_train)
                
                y_pred_train = model.predict(X_train)
                y_pred_test = model.predict(X_test)

                train_metrics.append(self.calculate_metrics(y_train, y_pred_train))
                test_metrics.append(self.calculate_metrics(y_test, y_pred_test))

            mean_train_metrics = np.mean(train_metrics, axis=0)
            mean_test_metrics = np.mean(test_metrics, axis=0)

            model_params = str(model.get_params())

            results = pd.DataFrame({
                'date': [pd.Timestamp.today().strftime('%Y-%m-%d')],
                'model_type': [model_name],
                'predictor_variables': [', '.join(X.columns)],  
                'model_parameters': [model_params], 
                'train_r2_mean': [mean_train_metrics[0]],
                'train_rmse_mean': [mean_train_metrics[1]],
                'train_mae_mean': [mean_train_metrics[2]],
                'train_mape_mean': [mean_train_metrics[3]],
                'train_me_mean': [mean_train_metrics[4]],
                'train_mpe_mean': [mean_train_metrics[5]],
                'test_r2_mean': [mean_test_metrics[0]],
                'test_rmse_mean': [mean_test_metrics[1]],
                'test_mae_mean': [mean_test_metrics[2]],
                'test_mape_mean': [mean_test_metrics[3]],
                'test_me_mean': [mean_test_metrics[4]],
                'test_mpe_mean': [mean_test_metrics[5]],
            })

            self.save_metrics_to_excel(results.to_dict(orient='records')[0])
            all_results.append(results)
        
        return all_results